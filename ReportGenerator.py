#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.styles import Font


class ReportGenerator:
    def __init__(self, env: str, run_id: str, dry_run: bool):
        self.env = env
        self.run_id = run_id
        self.dry_run = dry_run

        os.makedirs("reports", exist_ok=True)

        self.xlsx_filename = f"reports/autopatch_{env}_{run_id}.xlsx"
        self.json_filename = f"reports/autopatch_{env}_{run_id}.json"

    @staticmethod
    def _status_counts(items):
        ok = sum(1 for x in items if x.status == "OK")
        fail = sum(1 for x in items if x.status == "FAILED")
        skip = sum(1 for x in items if x.status == "SKIPPED")
        return ok, fail, skip, len(items)

    @staticmethod
    def _join_failed_hosts(value):
        if not value:
            return ""
        try:
            return ",".join(str(v) for v in value)
        except TypeError:
            return str(value)

    def _build_json_payload(self, standalone_outcomes, cluster_outcomes,
                            standalone_probe, cluster_probe):

        s_index = {o.host: o for o in standalone_outcomes}

        standalone_list = []
        for row in standalone_probe:
            probe = row["result"]
            outcome = s_index.get(row["host"])
            if not outcome:
                continue
            standalone_list.append({
                "host": probe.host,
                "probe": {
                    "ping_ok": bool(getattr(probe, "ping_ok", None)),
                    "ssh_ok": bool(getattr(probe, "ssh_ok", None)),
                    "ssh_login_ok": getattr(probe, "ssh_login_ok", None),
                    "used_user": getattr(probe, "used_user", None),
                    "autopatch_enabled": bool(row.get("autopatch_enabled", True)),
                    "freeipa_managed": bool(row.get("freeipa", False)),
                },
                "patch": {
                    "status": outcome.status,
                    "reason": outcome.reason,
                    "duration": outcome.duration,
                    "failed_hosts": list(getattr(outcome, "failed_hosts", []) or []),
                },
            })

        cluster_summary = []
        for co in cluster_outcomes:
            batch_list = []
            for b in (co.batch_results or []):
                user, dur, fh = b
                batch_list.append({
                    "user": user,
                    "duration": dur,
                    "failed_hosts": list(fh or []),
                })
            cluster_summary.append({
                "cluster": co.cluster,
                "status": co.status,
                "reason": co.reason,
                "duration_total": co.duration_total,
                "failed_hosts": list(co.failed_hosts or []),
                "batches": batch_list,
            })

        cluster_members = []
        for cname, rows in cluster_probe.items():
            for row in rows:
                probe = row["result"]
                cluster_members.append({
                    "cluster": cname,
                    "host": probe.host,
                    "ping_ok": bool(getattr(probe, "ping_ok", None)),
                    "ssh_ok": bool(getattr(probe, "ssh_ok", None)),
                    "ssh_login_ok": getattr(probe, "ssh_login_ok", None),
                    "used_user": getattr(probe, "used_user", None),
                    "autopatch_enabled": bool(row.get("autopatch_enabled", True)),
                    "freeipa_managed": bool(row.get("freeipa", False)),
                })

        payload = {
            "env": self.env,
            "run_id": self.run_id,
            "dry_run": self.dry_run,
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "standalone": {
                "items": standalone_list,
            },
            "clusters": {
                "summary": cluster_summary,
                "members": cluster_members,
            },
        }
        return payload

    def generate(self, standalone_outcomes, cluster_outcomes,
                 standalone_probe, cluster_probe):
        
        wb = Workbook()

        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = f"Autopatch rapport - {self.env} ({self.run_id})"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A3"] = "Typ"
        ws["B3"] = "OK"
        ws["C3"] = "FAILED"
        ws["D3"] = "SKIPPED"
        ws["E3"] = "Totalt"

        s_ok, s_fail, s_skip, s_total = self._status_counts(standalone_outcomes)
        c_ok, c_fail, c_skip, c_total = self._status_counts(cluster_outcomes)

        ws.append(["Standalone", s_ok, s_fail, s_skip, s_total])
        ws.append(["Kluster", c_ok, c_fail, c_skip, c_total])

        row = 8
        ws[f"A{row}"] = "Misslyckade standalone:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        for o in standalone_outcomes:
            if o.status == "FAILED":
                ws[f"A{row}"] = o.host
                row += 1

        row += 1
        ws[f"A{row}"] = "Misslyckade kluster:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        for o in cluster_outcomes:
            if o.status == "FAILED":
                ws[f"A{row}"] = o.cluster
                row += 1

        chart = PieChart()
        chart.title = "Standalone resultat"
        labels = Reference(ws, min_col=1, min_row=4, max_row=4)
        data = Reference(ws, min_col=2, min_row=4, max_col=4)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        ws.add_chart(chart, "G3")

        chart2 = PieChart()
        chart2.title = "Kluster resultat"
        labels2 = Reference(ws, min_col=1, min_row=5, max_row=5)
        data2 = Reference(ws, min_col=2, min_row=5, max_col=4)
        chart2.add_data(data2, titles_from_data=False)
        chart2.set_categories(labels2)
        ws.add_chart(chart2, "G18")

        ws2 = wb.create_sheet("Standalone")
        ws2.append([
            "Host", "Ping", "SSH", "Login",
            "Autopatch", "Status", "Duration",
            "Reason", "Failed hosts"
        ])

        s_index = {o.host: o for o in standalone_outcomes}

        for r in standalone_probe:
            outcome = s_index.get(r["host"])
            if not outcome:
                continue
            probe = r["result"]

            ws2.append([
                probe.host,
                bool(getattr(probe, "ping_ok", None)),
                bool(getattr(probe, "ssh_ok", None)),
                getattr(probe, "ssh_login_ok", None),
                bool(r.get("autopatch_enabled", True)),
                outcome.status,
                outcome.duration,
                outcome.reason,
                self._join_failed_hosts(getattr(outcome, "failed_hosts", None)),
            ])

        ws3 = wb.create_sheet("Clusters")
        ws3.append([
            "Cluster", "Status", "Duration total",
            "Reason", "Failed hosts", "Batch count"
        ])

        for co in cluster_outcomes:
            ws3.append([
                co.cluster,
                co.status,
                co.duration_total,
                co.reason,
                self._join_failed_hosts(co.failed_hosts),
                len(co.batch_results or []),
            ])

        ws4 = wb.create_sheet("ClusterMembers")
        ws4.append([
            "Cluster", "Host", "Ping", "SSH",
            "Login", "Autopatch"
        ])

        for cname, rows in cluster_probe.items():
            for r in rows:
                probe = r["result"]
                ws4.append([
                    cname,
                    probe.host,
                    bool(getattr(probe, "ping_ok", None)),
                    bool(getattr(probe, "ssh_ok", None)),
                    getattr(probe, "ssh_login_ok", None),
                    bool(r.get("autopatch_enabled", True)),
                ])

        wb.save(self.xlsx_filename)

        payload = self._build_json_payload(
            standalone_outcomes=standalone_outcomes,
            cluster_outcomes=cluster_outcomes,
            standalone_probe=standalone_probe,
            cluster_probe=cluster_probe,
        )
        with open(self.json_filename, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        return self.xlsx_filename, self.json_filename
